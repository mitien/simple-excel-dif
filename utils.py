from io import BytesIO
from typing import Tuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def load_excel_file(file) -> Optional[pd.DataFrame]:
    return pd.read_excel(file, keep_default_na=True, engine='openpyxl')


def compare_excels(df1: pd.DataFrame, df2: pd.DataFrame) -> Tuple[pd.DataFrame, bool]:
    orig_columns = df1.columns.tolist()

    df1 = df1.fillna("").astype(str)
    df2 = df2.fillna("").astype(str)

    if set(df1.columns) == set(df2.columns):
        # Use the original order from df1
        all_columns = orig_columns
        df2 = df2.reindex(columns=all_columns, fill_value="")
    else:
        # Use sorted union if columns differ
        all_columns = sorted(set(df1.columns).union(df2.columns))
        df1 = df1.reindex(columns=all_columns, fill_value="")
        df2 = df2.reindex(columns=all_columns, fill_value="")

    max_rows = max(len(df1), len(df2))
    df1 = df1.reindex(range(max_rows), fill_value="")
    df2 = df2.reindex(range(max_rows), fill_value="")

    diff = []
    for i in range(max_rows):
        row_diff = {}
        for col in all_columns:
            val1, val2 = df1.at[i, col], df2.at[i, col]
            if val1 != val2:
                row_diff[col] = f"{val1} → {val2}"
            else:
                row_diff[col] = val2
        diff.append(row_diff)

    result_df = pd.DataFrame(diff)
    # Always use the original order if columns match (ignoring order)
    if set(df1.columns) == set(df2.columns):
        result_df = result_df[orig_columns]
    else:
        result_df = result_df[all_columns]

    return result_df, True


def export_diff_to_excel(df) -> BytesIO:
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    # Write header
    for col_idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col)

    # Fill changes
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_idx, row in df.iterrows():
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            if "→" in str(val):
                cell.fill = yellow

    wb.save(output)
    output.seek(0)
    return output
